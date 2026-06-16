"""Rotas de importacao e consulta do cadastro CEAF.

Passo 15D.1 - Importador CEAF + tabela pacientes_ceaf.
"""

import csv
import hashlib
import io
import os
import re
import unicodedata
from datetime import date, datetime
from typing import Any, Dict, Iterable, List, Optional
from uuid import uuid4

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy import func, or_, text
from sqlalchemy.orm import Session

from database import engine
from models.consultorio_models import BaseConsultorio, PacienteCEAF, PacienteClinico, PacienteAgenda, ProntuarioClinico
from routers.consultorio import get_current_user_consultorio, get_db_consultorio

BaseConsultorio.metadata.create_all(bind=engine)


def _garantir_colunas_conversao_ceaf() -> None:
    """Adiciona colunas de vínculo CEAF de forma idempotente.

    O create_all cria tabelas novas, mas não altera tabelas já existentes no
    Supabase. Como pacientes_ceaf já pode estar criada na homologação, usamos
    ALTER TABLE de forma compatível com PostgreSQL e SQLite.
    """
    colunas = {
        "paciente_clinico_id": "INTEGER",
        "paciente_agenda_id": "INTEGER",
        "convertido_em": "TIMESTAMP",
        "conversao_status": "VARCHAR",
        "conversao_observacao": "TEXT",
    }
    try:
        with engine.begin() as conn:
            dialecto = engine.dialect.name
            existentes = set()
            if dialecto == "sqlite":
                resultado = conn.execute(text("PRAGMA table_info(pacientes_ceaf)"))
                existentes = {linha[1] for linha in resultado.fetchall()}
                for coluna, tipo in colunas.items():
                    if coluna not in existentes:
                        conn.execute(text(f"ALTER TABLE pacientes_ceaf ADD COLUMN {coluna} {tipo}"))
            else:
                for coluna, tipo in colunas.items():
                    conn.execute(text(f"ALTER TABLE pacientes_ceaf ADD COLUMN IF NOT EXISTS {coluna} {tipo}"))
    except Exception:
        # Não bloqueia inicialização; eventual inconsistência será revelada no
        # endpoint de conversão e poderá ser corrigida sem perda da base CEAF.
        pass


_garantir_colunas_conversao_ceaf()

router = APIRouter(prefix="/ceaf", tags=["CEAF - Pacientes"])

COLUNAS_ALIASES = {
    "cns": ["cns usuario", "cns usuário", "cns", "cartao sus", "cartão sus"],
    "cpf": ["cpf usuario", "cpf usuário", "cpf"],
    "nome": ["nome usuario", "nome usuário", "nome", "paciente", "nome paciente"],
    "medicamento_prescrito": ["medicamento prescrito", "medicamento", "medicamento solicitado"],
    "municipio": ["municipio usuario", "município usuário", "municipio", "município"],
    "logradouro": ["logradouro usuario", "logradouro usuário", "logradouro", "endereco", "endereço"],
    "numero_residencia": ["numero residencia", "número residência", "numero residência", "número residencia", "numero", "número"],
    "complemento_residencia": ["complemento residencia", "complemento residência", "complemento"],
    "data_fim_vigencia": ["data fim vigencia", "data fim vigência", "fim vigencia", "fim vigência", "vigencia final", "vigência final"],
    "situacao_lme": ["situacao lme", "situação lme", "situacao", "situação"],
    "data_inicio_medicamento": ["data inicio medicamento", "data início medicamento", "inicio medicamento", "início medicamento"],
    "telefone": ["fone paciente", "telefone", "telefone paciente"],
    "telefone_comercial": ["fone comerc. paciente", "fone comercial paciente", "telefone comercial"],
    "telefone_celular": ["fone cel. paciente", "fone celular paciente", "celular", "telefone celular"],
}


def _normalizar_texto(valor: Any) -> str:
    if valor is None:
        return ""
    texto = str(valor).strip()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"\s+", " ", texto)
    return texto.lower().strip()


def _somente_digitos(valor: Any) -> Optional[str]:
    texto = re.sub(r"\D+", "", str(valor or ""))
    return texto or None


def _limpar_string(valor: Any) -> Optional[str]:
    if valor is None:
        return None
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    texto = str(valor).strip()
    if texto in {"", "None", "nan", "NaN", "NaT"}:
        return None
    return texto


def _parse_data(valor: Any, datemode: int = 0) -> Optional[date]:
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, (int, float)):
        try:
            import xlrd

            return xlrd.xldate_as_datetime(valor, datemode).date()
        except Exception:
            return None
    texto = str(valor).strip()
    if not texto:
        return None
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(texto[:10], formato).date()
        except ValueError:
            continue
    return None


def _mapear_indices(cabecalho: List[Any]) -> Dict[str, int]:
    normalizados = [_normalizar_texto(c) for c in cabecalho]
    indices: Dict[str, int] = {}
    for campo, aliases in COLUNAS_ALIASES.items():
        aliases_norm = {_normalizar_texto(alias) for alias in aliases}
        for idx, nome_coluna in enumerate(normalizados):
            if nome_coluna in aliases_norm:
                indices[campo] = idx
                break
    return indices


def _gerar_chave(linha: Dict[str, Any]) -> str:
    base = "|".join(
        [
            linha.get("cpf") or "",
            linha.get("cns") or "",
            _normalizar_texto(linha.get("nome")),
            _normalizar_texto(linha.get("medicamento_prescrito")),
            str(linha.get("data_inicio_medicamento") or ""),
            str(linha.get("data_fim_vigencia") or ""),
        ]
    )
    return hashlib.sha1(base.encode("utf-8")).hexdigest()


def _linha_para_dict(linha: List[Any], indices: Dict[str, int], datemode: int = 0) -> Dict[str, Any]:
    def obter(campo: str) -> Any:
        idx = indices.get(campo)
        if idx is None or idx >= len(linha):
            return None
        return linha[idx]

    dados = {
        "cns": _somente_digitos(obter("cns")),
        "cpf": _somente_digitos(obter("cpf")),
        "nome": _limpar_string(obter("nome")),
        "medicamento_prescrito": _limpar_string(obter("medicamento_prescrito")),
        "municipio": _limpar_string(obter("municipio")),
        "logradouro": _limpar_string(obter("logradouro")),
        "numero_residencia": _limpar_string(obter("numero_residencia")),
        "complemento_residencia": _limpar_string(obter("complemento_residencia")),
        "data_fim_vigencia": _parse_data(obter("data_fim_vigencia"), datemode),
        "situacao_lme": _limpar_string(obter("situacao_lme")),
        "data_inicio_medicamento": _parse_data(obter("data_inicio_medicamento"), datemode),
        "telefone": _somente_digitos(obter("telefone")),
        "telefone_comercial": _somente_digitos(obter("telefone_comercial")),
        "telefone_celular": _somente_digitos(obter("telefone_celular")),
    }
    dados["chave_importacao"] = _gerar_chave(dados)
    return dados


def _ler_csv(conteudo: bytes) -> List[List[Any]]:
    texto = conteudo.decode("utf-8-sig", errors="replace")
    amostra = texto[:2048]
    dialect = csv.Sniffer().sniff(amostra, delimiters=";,\t,")
    return [list(row) for row in csv.reader(io.StringIO(texto), dialect)]


def _ler_xlsx(conteudo: bytes) -> List[List[Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _ler_xls(conteudo: bytes) -> tuple[List[List[Any]], int]:
    try:
        import xlrd
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="Dependência xlrd não instalada. Execute pip install xlrd==2.0.1 e refaça o deploy.",
        ) from exc

    book = xlrd.open_workbook(file_contents=conteudo)
    sheet = book.sheet_by_index(0)
    linhas = [sheet.row_values(i) for i in range(sheet.nrows)]
    return linhas, book.datemode


def _extrair_linhas(arquivo_nome: str, conteudo: bytes) -> tuple[List[List[Any]], int]:
    nome = (arquivo_nome or "").lower()
    if nome.endswith(".xls"):
        return _ler_xls(conteudo)
    if nome.endswith(".xlsx"):
        return _ler_xlsx(conteudo), 0
    if nome.endswith(".csv") or nome.endswith(".txt"):
        return _ler_csv(conteudo), 0
    raise HTTPException(status_code=400, detail="Formato não suportado. Use .xls, .xlsx ou .csv.")


def _preparar_registros(linhas: List[List[Any]], datemode: int) -> List[Dict[str, Any]]:
    if not linhas:
        return []
    cabecalho_idx = None
    indices = {}
    for idx, linha in enumerate(linhas[:10]):
        indices_tentativa = _mapear_indices(linha)
        if {"nome", "cpf"}.issubset(indices_tentativa) or {"nome", "cns"}.issubset(indices_tentativa):
            cabecalho_idx = idx
            indices = indices_tentativa
            break
    if cabecalho_idx is None:
        raise HTTPException(status_code=400, detail="Cabeçalho da planilha CEAF não identificado.")

    obrigatorios = ["nome"]
    faltantes = [campo for campo in obrigatorios if campo not in indices]
    if faltantes:
        raise HTTPException(status_code=400, detail=f"Colunas obrigatórias ausentes: {', '.join(faltantes)}")

    registros = []
    for linha in linhas[cabecalho_idx + 1 :]:
        if not any(_limpar_string(valor) for valor in linha):
            continue
        dados = _linha_para_dict(linha, indices, datemode)
        if not dados.get("nome"):
            continue
        registros.append(dados)
    return registros


def _atualizar_existente(paciente: PacienteCEAF, dados: Dict[str, Any], lote: str, origem: str) -> None:
    for campo, valor in dados.items():
        if campo == "chave_importacao":
            continue
        setattr(paciente, campo, valor)
    paciente.lote_importacao = lote
    paciente.origem_arquivo = origem
    paciente.ativo = True
    paciente.atualizado_em = datetime.utcnow()


@router.post("/pacientes/importar-planilha")
def importar_pacientes_ceaf(
    arquivo: UploadFile = File(...),
    atualizar_existentes: bool = Query(True, description="Atualiza dados já existentes pela chave de importação."),
    db: Session = Depends(get_db_consultorio),
    current=Depends(get_current_user_consultorio),
):
    conteudo = arquivo.file.read()
    if not conteudo:
        raise HTTPException(status_code=400, detail="Arquivo vazio.")

    linhas, datemode = _extrair_linhas(arquivo.filename, conteudo)
    registros = _preparar_registros(linhas, datemode)
    lote = f"ceaf-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{uuid4().hex[:8]}"

    inseridos = 0
    atualizados = 0
    ignorados = 0
    erros: List[Dict[str, Any]] = []

    for posicao, dados in enumerate(registros, start=1):
        try:
            existente = db.query(PacienteCEAF).filter(
                PacienteCEAF.chave_importacao == dados["chave_importacao"]
            ).first()
            if existente:
                if atualizar_existentes:
                    _atualizar_existente(existente, dados, lote, arquivo.filename)
                    atualizados += 1
                else:
                    ignorados += 1
                continue

            paciente = PacienteCEAF(
                **dados,
                lote_importacao=lote,
                origem_arquivo=arquivo.filename,
                ativo=True,
                criado_em=datetime.utcnow(),
                atualizado_em=datetime.utcnow(),
            )
            db.add(paciente)
            inseridos += 1
        except Exception as exc:
            erros.append({"linha": posicao, "erro": str(exc)[:300]})

    if erros:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail={"mensagem": "Falha ao preparar importação CEAF.", "erros": erros[:20]},
        )

    db.commit()

    return {
        "ok": True,
        "lote_importacao": lote,
        "arquivo": arquivo.filename,
        "total_linhas_validas": len(registros),
        "inseridos": inseridos,
        "atualizados": atualizados,
        "ignorados": ignorados,
    }


@router.get("/pacientes")
def listar_pacientes_ceaf(
    termo: Optional[str] = None,
    medicamento: Optional[str] = None,
    municipio: Optional[str] = None,
    situacao_lme: Optional[str] = None,
    vigencia_ate: Optional[date] = None,
    ativo: bool = True,
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db_consultorio),
    current=Depends(get_current_user_consultorio),
):
    query = db.query(PacienteCEAF).filter(PacienteCEAF.ativo == ativo)

    if termo:
        like = f"%{termo.strip()}%"
        query = query.filter(
            or_(
                PacienteCEAF.nome.ilike(like),
                PacienteCEAF.cpf.ilike(like),
                PacienteCEAF.cns.ilike(like),
                PacienteCEAF.telefone.ilike(like),
                PacienteCEAF.telefone_celular.ilike(like),
            )
        )
    if medicamento:
        query = query.filter(PacienteCEAF.medicamento_prescrito.ilike(f"%{medicamento.strip()}%"))
    if municipio:
        query = query.filter(PacienteCEAF.municipio.ilike(f"%{municipio.strip()}%"))
    if situacao_lme:
        query = query.filter(PacienteCEAF.situacao_lme.ilike(f"%{situacao_lme.strip()}%"))
    if vigencia_ate:
        query = query.filter(PacienteCEAF.data_fim_vigencia <= vigencia_ate)

    total = query.count()
    pacientes = query.order_by(PacienteCEAF.nome.asc()).offset(offset).limit(limit).all()

    return {"total": total, "limit": limit, "offset": offset, "pacientes": pacientes}


@router.get("/pacientes/resumo")
def resumo_ceaf(
    db: Session = Depends(get_db_consultorio),
    current=Depends(get_current_user_consultorio),
):
    total = db.query(func.count(PacienteCEAF.id)).scalar() or 0
    ativos = db.query(func.count(PacienteCEAF.id)).filter(PacienteCEAF.ativo.is_(True)).scalar() or 0
    com_cpf = db.query(func.count(PacienteCEAF.id)).filter(PacienteCEAF.cpf.isnot(None)).scalar() or 0
    com_cns = db.query(func.count(PacienteCEAF.id)).filter(PacienteCEAF.cns.isnot(None)).scalar() or 0
    medicamentos_distintos = db.query(func.count(func.distinct(PacienteCEAF.medicamento_prescrito))).scalar() or 0
    municipios_distintos = db.query(func.count(func.distinct(PacienteCEAF.municipio))).scalar() or 0
    proxima_vigencia = db.query(func.min(PacienteCEAF.data_fim_vigencia)).filter(
        PacienteCEAF.data_fim_vigencia.isnot(None),
        PacienteCEAF.ativo.is_(True),
    ).scalar()

    return {
        "total": total,
        "ativos": ativos,
        "com_cpf": com_cpf,
        "com_cns": com_cns,
        "medicamentos_distintos": medicamentos_distintos,
        "municipios_distintos": municipios_distintos,
        "proxima_vigencia": proxima_vigencia,
    }


def _telefone_preferencial(paciente: PacienteCEAF) -> Optional[str]:
    return paciente.telefone_celular or paciente.telefone or paciente.telefone_comercial


def _endereco_ceaf(paciente: PacienteCEAF) -> Optional[str]:
    partes = [paciente.logradouro, paciente.numero_residencia, paciente.complemento_residencia, paciente.municipio]
    texto = ", ".join([p for p in partes if p])
    return texto or None


def _query_pacientes_ceaf_para_conversao(
    db: Session,
    termo: Optional[str] = None,
    medicamento: Optional[str] = None,
    municipio: Optional[str] = None,
    situacao_lme: Optional[str] = None,
    vigencia_ate: Optional[date] = None,
    apenas_nao_convertidos: bool = True,
):
    query = db.query(PacienteCEAF).filter(PacienteCEAF.ativo.is_(True))

    if apenas_nao_convertidos:
        query = query.filter(PacienteCEAF.paciente_clinico_id.is_(None))

    if termo:
        like = f"%{termo.strip()}%"
        query = query.filter(
            or_(
                PacienteCEAF.nome.ilike(like),
                PacienteCEAF.cpf.ilike(like),
                PacienteCEAF.cns.ilike(like),
                PacienteCEAF.telefone.ilike(like),
                PacienteCEAF.telefone_celular.ilike(like),
            )
        )
    if medicamento:
        query = query.filter(PacienteCEAF.medicamento_prescrito.ilike(f"%{medicamento.strip()}%"))
    if municipio:
        query = query.filter(PacienteCEAF.municipio.ilike(f"%{municipio.strip()}%"))
    if situacao_lme:
        query = query.filter(PacienteCEAF.situacao_lme.ilike(f"%{situacao_lme.strip()}%"))
    if vigencia_ate:
        query = query.filter(PacienteCEAF.data_fim_vigencia <= vigencia_ate)

    return query


def _localizar_paciente_clinico(db: Session, paciente_ceaf: PacienteCEAF) -> Optional[PacienteClinico]:
    if paciente_ceaf.paciente_clinico_id:
        existente = db.query(PacienteClinico).filter(PacienteClinico.id == paciente_ceaf.paciente_clinico_id).first()
        if existente:
            return existente

    if paciente_ceaf.cpf:
        existente = db.query(PacienteClinico).filter(PacienteClinico.cpf == paciente_ceaf.cpf).first()
        if existente:
            return existente

    if paciente_ceaf.cns:
        existente = db.query(PacienteClinico).filter(PacienteClinico.cns == paciente_ceaf.cns).first()
        if existente:
            return existente

    return None


def _localizar_paciente_agenda(db: Session, paciente_ceaf: PacienteCEAF) -> Optional[PacienteAgenda]:
    if paciente_ceaf.paciente_agenda_id:
        existente = db.query(PacienteAgenda).filter(PacienteAgenda.id == paciente_ceaf.paciente_agenda_id).first()
        if existente:
            return existente

    if paciente_ceaf.cpf:
        existente = db.query(PacienteAgenda).filter(PacienteAgenda.cpf == paciente_ceaf.cpf).first()
        if existente:
            return existente

    if paciente_ceaf.cns:
        existente = db.query(PacienteAgenda).filter(PacienteAgenda.cns == paciente_ceaf.cns).first()
        if existente:
            return existente

    return None


def _criar_ou_vincular_paciente_agenda(db: Session, paciente_ceaf: PacienteCEAF) -> tuple[PacienteAgenda, str]:
    existente = _localizar_paciente_agenda(db, paciente_ceaf)
    telefone = _telefone_preferencial(paciente_ceaf)
    if existente:
        if telefone and not existente.telefone:
            existente.telefone = telefone
        if paciente_ceaf.cpf and not existente.cpf:
            existente.cpf = paciente_ceaf.cpf
        if paciente_ceaf.cns and not existente.cns:
            existente.cns = paciente_ceaf.cns
        if paciente_ceaf.municipio and not existente.municipio:
            existente.municipio = paciente_ceaf.municipio
        if paciente_ceaf.logradouro and not existente.logradouro:
            existente.logradouro = paciente_ceaf.logradouro
        if paciente_ceaf.numero_residencia and not existente.numero_residencia:
            existente.numero_residencia = paciente_ceaf.numero_residencia
        if paciente_ceaf.complemento_residencia and not existente.complemento_residencia:
            existente.complemento_residencia = paciente_ceaf.complemento_residencia
        existente.atualizado_em = datetime.utcnow()
        db.flush()
        return existente, "vinculado"

    novo = PacienteAgenda(
        nome=paciente_ceaf.nome,
        cpf=paciente_ceaf.cpf,
        cns=paciente_ceaf.cns,
        telefone=telefone,
        telefone_alternativo=paciente_ceaf.telefone_comercial,
        municipio=paciente_ceaf.municipio,
        logradouro=paciente_ceaf.logradouro,
        numero_residencia=paciente_ceaf.numero_residencia,
        complemento_residencia=paciente_ceaf.complemento_residencia,
        origem="ceaf",
        ativo=True,
        criado_em=datetime.utcnow(),
        atualizado_em=datetime.utcnow(),
    )
    db.add(novo)
    db.flush()
    return novo, "criado"


def _criar_ou_vincular_paciente_clinico(db: Session, paciente_ceaf: PacienteCEAF) -> tuple[PacienteClinico, str]:
    existente = _localizar_paciente_clinico(db, paciente_ceaf)
    telefone = _telefone_preferencial(paciente_ceaf)
    endereco = _endereco_ceaf(paciente_ceaf)
    observacao_ceaf = (
        "Origem CEAF. "
        f"Medicamento prescrito: {paciente_ceaf.medicamento_prescrito or 'não informado'}. "
        f"Situação LME: {paciente_ceaf.situacao_lme or 'não informada'}. "
        f"Vigência até: {paciente_ceaf.data_fim_vigencia or 'não informada'}."
    )

    if existente:
        if telefone and not existente.telefone:
            existente.telefone = telefone
        if endereco and not existente.endereco:
            existente.endereco = endereco
        if paciente_ceaf.cpf and not existente.cpf:
            existente.cpf = paciente_ceaf.cpf
        if paciente_ceaf.cns and not existente.cns:
            existente.cns = paciente_ceaf.cns
        if not existente.origem or existente.origem == "conversao_servico_rapido":
            existente.origem = "ceaf"
        if observacao_ceaf and not existente.observacoes_clinicas:
            existente.observacoes_clinicas = observacao_ceaf
        db.flush()
        return existente, "vinculado"

    novo = PacienteClinico(
        nome=paciente_ceaf.nome,
        telefone=telefone,
        endereco=endereco,
        cpf=paciente_ceaf.cpf,
        cns=paciente_ceaf.cns,
        origem="ceaf",
        aceite_verbal=True,
        motivo_conversao="Conversão em lote a partir do cadastro CEAF para uso em agenda, notificações e cuidado farmacêutico.",
        observacoes_clinicas=observacao_ceaf,
        criado_em=datetime.utcnow(),
    )
    db.add(novo)
    db.flush()
    return novo, "criado"


def _garantir_prontuario(db: Session, paciente: PacienteClinico) -> tuple[ProntuarioClinico, str]:
    prontuario = db.query(ProntuarioClinico).filter(ProntuarioClinico.paciente_clinico_id == paciente.id).first()
    if prontuario:
        return prontuario, "existente"
    prontuario = ProntuarioClinico(
        paciente_clinico_id=paciente.id,
        status="ativo",
        data_abertura=datetime.utcnow(),
        observacoes="Prontuário aberto automaticamente a partir da conversão CEAF.",
    )
    db.add(prontuario)
    db.flush()
    return prontuario, "criado"


def _converter_paciente_ceaf(db: Session, paciente_ceaf: PacienteCEAF) -> Dict[str, Any]:
    if not paciente_ceaf.cpf and not paciente_ceaf.cns:
        paciente_ceaf.conversao_status = "ignorado_sem_identificador"
        paciente_ceaf.conversao_observacao = "Registro sem CPF e CNS; conversão automática não realizada."
        paciente_ceaf.atualizado_em = datetime.utcnow()
        return {"status": "ignorado", "motivo": "sem_cpf_cns", "ceaf_id": paciente_ceaf.id}

    paciente_agenda, agenda_status = _criar_ou_vincular_paciente_agenda(db, paciente_ceaf)
    paciente_clinico, clinico_status = _criar_ou_vincular_paciente_clinico(db, paciente_ceaf)
    prontuario, prontuario_status = _garantir_prontuario(db, paciente_clinico)

    paciente_ceaf.paciente_agenda_id = paciente_agenda.id
    paciente_ceaf.paciente_clinico_id = paciente_clinico.id
    paciente_ceaf.convertido_em = datetime.utcnow()
    paciente_ceaf.conversao_status = "convertido"
    paciente_ceaf.conversao_observacao = (
        f"Paciente agenda {agenda_status}; paciente clínico {clinico_status}; prontuário {prontuario_status}."
    )
    paciente_ceaf.atualizado_em = datetime.utcnow()

    return {
        "status": "convertido",
        "ceaf_id": paciente_ceaf.id,
        "paciente_agenda_id": paciente_agenda.id,
        "paciente_clinico_id": paciente_clinico.id,
        "prontuario_id": prontuario.id,
        "agenda_status": agenda_status,
        "clinico_status": clinico_status,
        "prontuario_status": prontuario_status,
    }


@router.get("/pacientes/conversao/resumo")
def resumo_conversao_ceaf(
    db: Session = Depends(get_db_consultorio),
    current=Depends(get_current_user_consultorio),
):
    total = db.query(func.count(PacienteCEAF.id)).filter(PacienteCEAF.ativo.is_(True)).scalar() or 0
    convertidos = db.query(func.count(PacienteCEAF.id)).filter(
        PacienteCEAF.ativo.is_(True),
        PacienteCEAF.paciente_clinico_id.isnot(None),
    ).scalar() or 0
    sem_identificador = db.query(func.count(PacienteCEAF.id)).filter(
        PacienteCEAF.ativo.is_(True),
        PacienteCEAF.cpf.is_(None),
        PacienteCEAF.cns.is_(None),
    ).scalar() or 0
    pendentes = max(total - convertidos - sem_identificador, 0)

    return {
        "total_ceaf_ativos": total,
        "convertidos": convertidos,
        "pendentes_conversao": pendentes,
        "sem_cpf_cns": sem_identificador,
    }


@router.post("/pacientes/{paciente_id}/converter-clinico")
def converter_paciente_ceaf_individual(
    paciente_id: int,
    db: Session = Depends(get_db_consultorio),
    current=Depends(get_current_user_consultorio),
):
    paciente = db.query(PacienteCEAF).filter(PacienteCEAF.id == paciente_id).first()
    if not paciente:
        raise HTTPException(status_code=404, detail="Paciente CEAF não encontrado.")

    resultado = _converter_paciente_ceaf(db, paciente)
    db.commit()
    return {"ok": True, "resultado": resultado}


@router.post("/pacientes/converter-lote")
def converter_pacientes_ceaf_lote(
    termo: Optional[str] = None,
    medicamento: Optional[str] = None,
    municipio: Optional[str] = None,
    situacao_lme: Optional[str] = None,
    vigencia_ate: Optional[date] = None,
    apenas_nao_convertidos: bool = True,
    limite: int = Query(1000, ge=1, le=5000),
    db: Session = Depends(get_db_consultorio),
    current=Depends(get_current_user_consultorio),
):
    query = _query_pacientes_ceaf_para_conversao(
        db=db,
        termo=termo,
        medicamento=medicamento,
        municipio=municipio,
        situacao_lme=situacao_lme,
        vigencia_ate=vigencia_ate,
        apenas_nao_convertidos=apenas_nao_convertidos,
    )

    pacientes = query.order_by(PacienteCEAF.id.asc()).limit(limite).all()
    resultados: List[Dict[str, Any]] = []
    criados_clinicos = 0
    vinculados_clinicos = 0
    criados_agenda = 0
    vinculados_agenda = 0
    prontuarios_criados = 0
    ignorados = 0
    erros: List[Dict[str, Any]] = []

    for paciente in pacientes:
        try:
            resultado = _converter_paciente_ceaf(db, paciente)
            resultados.append(resultado)
            if resultado.get("status") == "ignorado":
                ignorados += 1
                continue
            if resultado.get("clinico_status") == "criado":
                criados_clinicos += 1
            if resultado.get("clinico_status") == "vinculado":
                vinculados_clinicos += 1
            if resultado.get("agenda_status") == "criado":
                criados_agenda += 1
            if resultado.get("agenda_status") == "vinculado":
                vinculados_agenda += 1
            if resultado.get("prontuario_status") == "criado":
                prontuarios_criados += 1
        except Exception as exc:
            db.rollback()
            erros.append({"ceaf_id": paciente.id, "nome": paciente.nome, "erro": str(exc)[:300]})

    if erros:
        raise HTTPException(
            status_code=400,
            detail={"mensagem": "Falha durante conversão em lote CEAF.", "erros": erros[:20]},
        )

    db.commit()

    return {
        "ok": True,
        "total_analisados": len(pacientes),
        "convertidos_ou_vinculados": len([r for r in resultados if r.get("status") == "convertido"]),
        "novos_pacientes_clinicos": criados_clinicos,
        "pacientes_clinicos_vinculados": vinculados_clinicos,
        "novos_pacientes_agenda": criados_agenda,
        "pacientes_agenda_vinculados": vinculados_agenda,
        "prontuarios_criados": prontuarios_criados,
        "ignorados_sem_cpf_cns": ignorados,
        "limite": limite,
        "amostra_resultados": resultados[:20],
    }


@router.get("/pacientes/{paciente_id}")
def obter_paciente_ceaf(
    paciente_id: int,
    db: Session = Depends(get_db_consultorio),
    current=Depends(get_current_user_consultorio),
):
    paciente = db.query(PacienteCEAF).filter(PacienteCEAF.id == paciente_id).first()
    if not paciente:
        raise HTTPException(status_code=404, detail="Paciente CEAF não encontrado.")
    return {"paciente": paciente}
