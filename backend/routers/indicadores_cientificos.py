from io import BytesIO
from datetime import datetime
from collections import defaultdict

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from models.consultorio_models import (
    PacienteSimplificado,
    AtendimentoRapido,
    AfericaoPA,
    GlicemiaCapilar,
    Bioimpedancia,
    PicoFluxo,
    PacienteClinico,
    IntervencaoFarmacoterapia,
    DesfechoIntervencaoFarmacoterapia,
    PlanoCuidado,
)
from routers.consultorio import get_db_consultorio, get_current_user_consultorio
from services.consultorio_helpers import calcular_percentual

router = APIRouter(
    prefix="/consultorio",
    tags=["Consultório Farmacêutico - Indicadores Científicos"]
)

@router.get("/dashboard-epidemiologico")
def dashboard_epidemiologico(
    db: Session = Depends(get_db_consultorio),
    current=Depends(get_current_user_consultorio)
):
    pacientes = db.query(PacienteClinico).all()

    total = len(pacientes)

    if total == 0:
        return {
            "total_pacientes": 0,
            "sexo": {},
            "faixa_etaria": {},
            "principais_cids": {},
            "bairros": {}
        }

    sexo = {}
    faixa_etaria = {}
    principais_cids = {}
    bairros = {}

    for p in pacientes:

        # SEXO
        sexo_key = p.sexo or "Não informado"
        sexo[sexo_key] = sexo.get(sexo_key, 0) + 1

        # IDADE
        idade = p.idade or 0

        if idade < 12:
            faixa = "0-11"
        elif idade < 18:
            faixa = "12-17"
        elif idade < 40:
            faixa = "18-39"
        elif idade < 60:
            faixa = "40-59"
        else:
            faixa = "60+"

        faixa_etaria[faixa] = faixa_etaria.get(faixa, 0) + 1

        # CID
        cid = p.cid_principal or "Não informado"
        principais_cids[cid] = principais_cids.get(cid, 0) + 1

        # BAIRRO
        bairro = p.bairro or "Não informado"
        bairros[bairro] = bairros.get(bairro, 0) + 1

    principais_cids = dict(
        sorted(
            principais_cids.items(),
            key=lambda x: x[1],
            reverse=True
        )[:10]
    )

    bairros = dict(
        sorted(
            bairros.items(),
            key=lambda x: x[1],
            reverse=True
        )[:10]
    )

    return {
        "total_pacientes": total,
        "sexo": sexo,
        "faixa_etaria": faixa_etaria,
        "principais_cids": principais_cids,
        "bairros": bairros
    }

@router.get("/dashboard-antropometrico")
def dashboard_antropometrico(
    db: Session = Depends(get_db_consultorio),
    current=Depends(get_current_user_consultorio)
):
    registros = db.query(Bioimpedancia).all()

    total = len(registros)

    if total == 0:
        return {
            "total_avaliacoes": 0,
            "media_imc": 0,
            "media_peso": 0,
            "media_gordura_corporal": 0,
            "media_massa_muscular": 0,
            "media_gordura_visceral": 0,
            "classificacoes_imc": {},
            "classificacoes_gordura_visceral": {}
        }

    imcs = []
    pesos = []
    gorduras = []
    massas = []
    viscerais = []

    classificacoes_imc = {}
    classificacoes_gordura_visceral = {}

    for r in registros:
        if getattr(r, "imc", None) is not None:
            imcs.append(r.imc)

        if getattr(r, "peso", None) is not None:
            pesos.append(r.peso)

        if getattr(r, "percentual_gordura", None) is not None:
            gorduras.append(r.percentual_gordura)

        if getattr(r, "percentual_massa_muscular", None) is not None:
            massas.append(r.percentual_massa_muscular)

        if getattr(r, "gordura_visceral", None) is not None:
            viscerais.append(r.gordura_visceral)

        classe_imc = getattr(r, "classificacao_imc", None) or "não_classificado"
        classificacoes_imc[classe_imc] = classificacoes_imc.get(classe_imc, 0) + 1

        classe_visceral = (
            getattr(r, "classificacao_gordura_visceral", None)
            or "não_classificado"
        )

        classificacoes_gordura_visceral[classe_visceral] = (
            classificacoes_gordura_visceral.get(classe_visceral, 0) + 1
        )

    return {
        "total_avaliacoes": total,
        "media_imc": round(sum(imcs) / len(imcs), 2) if imcs else 0,
        "media_peso": round(sum(pesos) / len(pesos), 2) if pesos else 0,
        "media_gordura_corporal": round(sum(gorduras) / len(gorduras), 2) if gorduras else 0,
        "media_massa_muscular": round(sum(massas) / len(massas), 2) if massas else 0,
        "media_gordura_visceral": round(sum(viscerais) / len(viscerais), 2) if viscerais else 0,
        "classificacoes_imc": classificacoes_imc,
        "classificacoes_gordura_visceral": classificacoes_gordura_visceral
    }

@router.get("/dashboard-cardiovascular")
def dashboard_cardiovascular(
    db: Session = Depends(get_db_consultorio),
    current=Depends(get_current_user_consultorio)
):
    afericoes = db.query(AfericaoPA).all()

    total = len(afericoes)

    if total == 0:
        return {
            "total_afericoes": 0,
            "media_pas": 0,
            "media_pad": 0,
            "classificacoes": {}
        }

    soma_pas = 0
    soma_pad = 0

    classificacoes = {
        "normal": 0,
        "pa_elevada": 0,
        "hipertensao": 0,
        "crise_hipertensiva": 0,
    }

    for a in afericoes:
        soma_pas += a.pressao_sistolica or 0
        soma_pad += a.pressao_diastolica or 0

        classe = a.classificacao or "não_classificado"
        classificacoes[classe] = classificacoes.get(classe, 0) + 1

    return {
        "total_afericoes": total,
        "media_pas": round(soma_pas / total, 2),
        "media_pad": round(soma_pad / total, 2),
        "classificacoes": classificacoes
    }

@router.get("/dashboard-glicemico")
def dashboard_glicemico(
    db: Session = Depends(get_db_consultorio),
    current=Depends(get_current_user_consultorio)
):
    registros = db.query(GlicemiaCapilar).all()

    total = len(registros)

    if total == 0:
        return {
            "total_afericoes": 0,
            "media_glicemia": 0,
            "classificacoes": {},
            "tipos_jejum": {}
        }

    valores = []
    classificacoes = {}
    tipos_jejum = {}

    for r in registros:
        if getattr(r, "valor_glicemia", None) is not None:
            valores.append(r.valor_glicemia)

        classe = getattr(r, "classificacao", None) or "não_classificado"
        classificacoes[classe] = classificacoes.get(classe, 0) + 1

        tipo = getattr(r, "tipo_jejum", None) or "não_informado"
        tipos_jejum[tipo] = tipos_jejum.get(tipo, 0) + 1

    return {
        "total_afericoes": total,
        "media_glicemia": round(sum(valores) / len(valores), 2) if valores else 0,
        "classificacoes": classificacoes,
        "tipos_jejum": tipos_jejum
    }

@router.get("/dashboard-efetividade-cuidado")
def dashboard_efetividade_cuidado(
    db: Session = Depends(get_db_consultorio),
    current=Depends(get_current_user_consultorio)
):
    planos = db.query(PlanoCuidado).all()

    total = len(planos)

    ativos = len([
        p for p in planos
        if p.status in ["pendente", "em_acompanhamento"]
    ])

    concluidos = len([
        p for p in planos
        if p.status == "concluido"
    ])

    cancelados = len([
        p for p in planos
        if p.status == "cancelado"
    ])

    atingidos = len([
        p for p in planos
        if p.resultado_classificacao == "atingido"
    ])

    parcialmente = len([
        p for p in planos
        if p.resultado_classificacao == "parcialmente_atingido"
    ])

    nao_atingidos = len([
        p for p in planos
        if p.resultado_classificacao == "nao_atingido"
    ])

    taxa_sucesso = (
        round((atingidos / concluidos) * 100, 2)
        if concluidos > 0
        else 0
    )

    distribuicao_problemas = {}

    for plano in planos:
        problema = (
            plano.problema_identificado
            or "Não informado"
        )

        distribuicao_problemas[problema] = (
            distribuicao_problemas.get(problema, 0) + 1
        )

    tempo_medio = 0

    tempos = []

    for plano in planos:
        if plano.concluido_em and plano.criado_em:
            dias = (
                plano.concluido_em.date()
                - plano.criado_em.date()
            ).days

            tempos.append(dias)

    if tempos:
        tempo_medio = round(
            sum(tempos) / len(tempos),
            1
        )

    return {
        "planos": {
            "total": total,
            "ativos": ativos,
            "concluidos": concluidos,
            "cancelados": cancelados,
        },

        "resultados": {
            "atingidos": atingidos,
            "parcialmente_atingidos": parcialmente,
            "nao_atingidos": nao_atingidos,
        },

        "taxa_sucesso": taxa_sucesso,

        "tempo_medio_conclusao_dias": tempo_medio,

        "problemas": distribuicao_problemas,
    }

@router.get("/indicadores-cientificos")
def indicadores_cientificos(
    db: Session = Depends(get_db_consultorio),
    current=Depends(get_current_user_consultorio)
):
    pacientes = db.query(PacienteClinico).all()
    afericoes_pa = db.query(AfericaoPA).all()
    glicemias = db.query(GlicemiaCapilar).all()
    bioimpedancias = db.query(Bioimpedancia).all()
    intervencoes = db.query(IntervencaoFarmacoterapia).all()

    sexo = {}

    for p in pacientes:
        sexo_key = getattr(p, "sexo", None) or "Não informado"
        sexo[sexo_key] = sexo.get(sexo_key, 0) + 1

    pas = [
        getattr(a, "pressao_sistolica", None)
        for a in afericoes_pa
        if getattr(a, "pressao_sistolica", None) is not None
    ]

    pad = [
        getattr(a, "pressao_diastolica", None)
        for a in afericoes_pa
        if getattr(a, "pressao_diastolica", None) is not None
    ]

    glicemia_valores = [
        getattr(g, "valor_glicemia", None)
        for g in glicemias
        if getattr(g, "valor_glicemia", None) is not None
    ]

    imcs = [
        getattr(b, "imc", None)
        for b in bioimpedancias
        if getattr(b, "imc", None) is not None
    ]

    intervencoes_aceitas = sum(
        1 for i in intervencoes
        if getattr(i, "aceita_pelo_paciente", False)
    )

    encaminhamentos = sum(
        1 for i in intervencoes
        if getattr(i, "necessidade_encaminhamento", False)
    )

    return {
        "assistencial": {
            "total_pacientes_clinicos": len(pacientes),
            "total_afericoes_pa": len(afericoes_pa),
            "total_glicemias": len(glicemias),
            "total_bioimpedancias": len(bioimpedancias),
            "total_intervencoes": len(intervencoes),
        },
        "perfil_pacientes": {
            "sexo": sexo
        },
        "cardiovascular": {
            "media_pas": round(sum(pas) / len(pas), 2) if pas else 0,
            "media_pad": round(sum(pad) / len(pad), 2) if pad else 0,
        },
        "glicemico": {
            "media_glicemia": round(sum(glicemia_valores) / len(glicemia_valores), 2)
            if glicemia_valores else 0,
        },
        "antropometrico": {
            "media_imc": round(sum(imcs) / len(imcs), 2) if imcs else 0,
        },
        "intervencoes_farmaceuticas": {
            "intervencoes_aceitas": intervencoes_aceitas,
            "encaminhamentos": encaminhamentos,
            "taxa_aceitacao": round(
                (intervencoes_aceitas / len(intervencoes)) * 100,
                2
            ) if intervencoes else 0,
            "taxa_encaminhamento": round(
                (encaminhamentos / len(intervencoes)) * 100,
                2
            ) if intervencoes else 0,
        }
    }    
@router.get("/serie-temporal-cientifica")
def serie_temporal_cientifica(
    db: Session = Depends(get_db_consultorio),
    current=Depends(get_current_user_consultorio)
):
    afericoes_pa = db.query(AfericaoPA).all()
    glicemias = db.query(GlicemiaCapilar).all()
    bioimpedancias = db.query(Bioimpedancia).all()
    intervencoes = db.query(IntervencaoFarmacoterapia).all()

    series = defaultdict(lambda: {
        "afericoes_pa": 0,
        "glicemias": 0,
        "bioimpedancias": 0,
        "intervencoes": 0,
    })

    def obter_mes(data):
        if not data:
            return "Sem data"

        try:
            return data.strftime("%Y-%m")
        except:
            return "Sem data"

    for a in afericoes_pa:
        mes = obter_mes(getattr(a, "data_afericao", None))
        series[mes]["afericoes_pa"] += 1

    for g in glicemias:
        mes = obter_mes(getattr(g, "data_afericao", None))
        series[mes]["glicemias"] += 1

    for b in bioimpedancias:
        mes = obter_mes(getattr(b, "data_avaliacao", None))
        series[mes]["bioimpedancias"] += 1

    for i in intervencoes:
        mes = obter_mes(getattr(i, "created_at", None))
        series[mes]["intervencoes"] += 1

    resultado = []

    for mes in sorted(series.keys()):
        resultado.append({
            "mes": mes,
            **series[mes]
        })

    return resultado

@router.get("/exportacao-cientifica-excel")
def exportacao_cientifica_excel(
    db: Session = Depends(get_db_consultorio),
    current=Depends(get_current_user_consultorio)
):
    wb = Workbook()

    # ABA 1 — INDICADORES
    ws = wb.active
    ws.title = "Indicadores"

    indicadores = indicadores_cientificos(
        db=db,
        current=current
    )

    ws.append(["Categoria", "Indicador", "Valor"])

    for categoria, dados in indicadores.items():
        if isinstance(dados, dict):
            for indicador, valor in dados.items():
                if isinstance(valor, dict):
                    for subindicador, subvalor in valor.items():
                        ws.append([
                            categoria,
                            f"{indicador} - {subindicador}",
                            subvalor
                        ])
                else:
                    ws.append([
                        categoria,
                        indicador,
                        valor
                    ])

    # ABA 2 — SÉRIE TEMPORAL
    ws2 = wb.create_sheet("Serie_Temporal")

    serie = serie_temporal_cientifica(
        db=db,
        current=current
    )

    ws2.append([
        "Mês",
        "Aferições PA",
        "Glicemias",
        "Bioimpedâncias",
        "Intervenções"
    ])

    for item in serie:
        ws2.append([
            item.get("mes"),
            item.get("afericoes_pa"),
            item.get("glicemias"),
            item.get("bioimpedancias"),
            item.get("intervencoes")
        ])

    # ABA 3 — METADADOS
    ws3 = wb.create_sheet("Metadados")

    ws3.append(["Campo", "Valor"])
    ws3.append(["Data de exportação", datetime.now().strftime("%d/%m/%Y %H:%M")])
    ws3.append(["Profissional", getattr(current, "nome", "Não informado")])
    ws3.append(["Categoria profissional", getattr(current, "categoria_profissional", "Não informado")])
    ws3.append(["Finalidade", "Exportação científica agregada e anonimizada"])

    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            column_letter = column_cells[0].column_letter
            sheet.column_dimensions[column_letter].width = 28

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=exportacao_cientifica.xlsx"
        }
    )

@router.get("/exportacao-pesquisa-anonimizada")
def exportacao_pesquisa_anonimizada(
    db: Session = Depends(get_db_consultorio),
    current=Depends(get_current_user_consultorio)
):
    wb = Workbook()

    # ABA 1 — PACIENTES
    ws = wb.active
    ws.title = "Pacientes"

    ws.append([
        "codigo_paciente",
        "idade",
        "sexo",
        "bairro",
        "convertido_consultorio"
    ])

    pacientes = db.query(PacienteSimplificado).all()

    for p in pacientes:
        convertido = db.query(PacienteClinico).filter(
            PacienteClinico.paciente_simplificado_origem_id == p.id
        ).first()

        ws.append([
            f"P{p.id}",
            getattr(p, "idade", None),
            getattr(p, "sexo", None),
            getattr(p, "bairro", None),
            "SIM" if convertido else "NÃO"
        ])

    # ABA 2 — SERVIÇOS RÁPIDOS
    ws2 = wb.create_sheet("Servicos_Rapidos")

    ws2.append([
        "codigo_paciente",
        "codigo_atendimento",
        "data_atendimento",
        "tipo_servico",
        "pas",
        "pad",
        "frequencia_cardiaca",
        "classificacao_pa",
        "glicemia",
        "tipo_glicemia",
        "classificacao_glicemia",
        "pfe",
        "pfe_previsto",
        "pfe_percentual",
        "classificacao_pfe",
        "peso",
        "altura",
        "imc",
        "classificacao_imc",
        "gordura_corporal_percentual",
        "massa_muscular_percentual",
        "gordura_visceral",
        "classificacao_gordura_visceral",
        "fmi",
        "ffmi"
    ])

    atendimentos = db.query(AtendimentoRapido).all()

    for a in atendimentos:
        pa = db.query(AfericaoPA).filter(
            AfericaoPA.atendimento_rapido_id == a.id
        ).first()

        glicemia = db.query(GlicemiaCapilar).filter(
            GlicemiaCapilar.atendimento_rapido_id == a.id
        ).first()

        pico = db.query(PicoFluxo).filter(
            PicoFluxo.atendimento_rapido_id == a.id
        ).first()

        bio = db.query(Bioimpedancia).filter(
            Bioimpedancia.atendimento_rapido_id == a.id
        ).first()

        ws2.append([
            f"P{a.paciente_simplificado_id}",
            f"A{a.id}",
            getattr(a, "data_atendimento", None),
            getattr(a, "tipo_servico", None),

            getattr(pa, "pressao_sistolica", None) if pa else None,
            getattr(pa, "pressao_diastolica", None) if pa else None,
            getattr(pa, "frequencia_cardiaca", None) if pa else None,
            getattr(pa, "classificacao", None) if pa else None,

            getattr(glicemia, "valor_glicemia", None) if glicemia else None,
            getattr(glicemia, "tipo_jejum", None) if glicemia else None,
            getattr(glicemia, "classificacao", None) if glicemia else None,

            getattr(pico, "valor_medido", None) if pico else None,
            getattr(pico, "valor_previsto", None) if pico else None,
            getattr(pico, "percentual_previsto", None) if pico else None,
            getattr(pico, "classificacao", None) if pico else None,

            getattr(bio, "peso", None) if bio else None,
            getattr(bio, "altura", None) if bio else None,
            getattr(bio, "imc", None) if bio else None,
            getattr(bio, "classificacao_imc", None) if bio else None,
            getattr(bio, "percentual_gordura", None) if bio else None,
            getattr(bio, "percentual_massa_muscular", None) if bio else None,
            getattr(bio, "gordura_visceral", None) if bio else None,
            getattr(bio, "classificacao_gordura_visceral", None) if bio else None,
            getattr(bio, "fmi", None) if bio else None,
            getattr(bio, "ffmi", None) if bio else None,
        ])

    # ABA 3 — INTERVENÇÕES FARMACÊUTICAS
    ws3 = wb.create_sheet("Intervencoes")

    ws3.append([
        "codigo_paciente_clinico",
        "codigo_intervencao",
        "data_intervencao",
        "tipo_intervencao",
        "aceita_pelo_paciente",
        "necessidade_encaminhamento",
        "status_desfecho",
        "necessidade_nova_intervencao"
    ])

    intervencoes = db.query(IntervencaoFarmacoterapia).all()

    for i in intervencoes:
        desfecho = db.query(DesfechoIntervencaoFarmacoterapia).filter(
            DesfechoIntervencaoFarmacoterapia.intervencao_id == i.id
        ).order_by(
            DesfechoIntervencaoFarmacoterapia.criado_em.desc()
        ).first()

        ws3.append([
            f"PC{i.paciente_clinico_id}",
            f"IF{i.id}",
            getattr(i, "criado_em", None),
            getattr(i, "tipo_intervencao", None),
            getattr(i, "aceita_pelo_paciente", None),
            getattr(i, "necessidade_encaminhamento", None),
            getattr(desfecho, "status_desfecho", None) if desfecho else None,
            getattr(desfecho, "necessidade_nova_intervencao", None) if desfecho else None,
        ])

    # ABA 4 — DICIONÁRIO DE DADOS
    ws4 = wb.create_sheet("Dicionario_Dados")

    ws4.append(["variavel", "descricao"])

    dicionario = [
        ["codigo_paciente", "Identificador pseudonimizado do paciente simplificado"],
        ["codigo_paciente_clinico", "Identificador pseudonimizado do paciente clínico"],
        ["idade", "Idade registrada no cadastro"],
        ["sexo", "Sexo registrado"],
        ["bairro", "Bairro de residência"],
        ["pas", "Pressão arterial sistólica"],
        ["pad", "Pressão arterial diastólica"],
        ["glicemia", "Valor de glicemia capilar"],
        ["pfe", "Pico de fluxo expiratório medido"],
        ["imc", "Índice de massa corporal"],
        ["fmi", "Índice de massa gorda"],
        ["ffmi", "Índice de massa livre de gordura"],
        ["tipo_intervencao", "Tipo de intervenção farmacêutica registrada"],
        ["status_desfecho", "Desfecho da intervenção farmacêutica"],
    ]

    for linha in dicionario:
        ws4.append(linha)

    # ABA 5 — METADADOS
    ws5 = wb.create_sheet("Metadados")

    ws5.append(["Campo", "Valor"])
    ws5.append(["Data de exportação", datetime.now().strftime("%d/%m/%Y %H:%M")])
    ws5.append(["Finalidade", "Base anonimizada para pesquisa e análise epidemiológica"])
    ws5.append(["Campos removidos", "Nome, CPF, CNS, telefone, endereço e nome da mãe"])
    ws5.append(["Identificação", "Pseudonimizada por códigos internos não nominais"])

    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            column_letter = column_cells[0].column_letter
            sheet.column_dimensions[column_letter].width = 28

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            "attachment; filename=pesquisa_anonimizada_completa.xlsx"
        }
    )

@router.get("/relatorio-cientifico-pdf")
def relatorio_cientifico_pdf(
    db: Session = Depends(get_db_consultorio),
    current=Depends(get_current_user_consultorio)
):
    indicadores = indicadores_cientificos(
        db=db,
        current=current
    )

    serie = serie_temporal_cientifica(
        db=db,
        current=current
    )

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph("FARMÁCIA ESCOLA", styles["Title"]))
    elementos.append(Paragraph("PROFª ANA MARIA CERVANTES BARAZA", styles["Heading2"]))
    elementos.append(Paragraph("Universidade Federal de Mato Grosso do Sul", styles["Normal"]))
    elementos.append(Spacer(1, 14))

    elementos.append(Paragraph("Relatório Científico Automatizado", styles["Title"]))
    elementos.append(Paragraph(
        f"Emitido em: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles["Normal"]
    ))
    elementos.append(Spacer(1, 18))

    def adicionar_tabela(titulo, linhas):
        elementos.append(Paragraph(titulo, styles["Heading2"]))

        tabela = Table(linhas, colWidths=[260, 220])

        tabela.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e0f2f1")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("PADDING", (0, 0), (-1, -1), 7),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))

        elementos.append(tabela)
        elementos.append(Spacer(1, 16))

    assistencial = indicadores.get("assistencial", {})
    cardiovascular = indicadores.get("cardiovascular", {})
    glicemico = indicadores.get("glicemico", {})
    antropometrico = indicadores.get("antropometrico", {})
    intervencoes = indicadores.get("intervencoes_farmaceuticas", {})

    adicionar_tabela("Indicadores Assistenciais", [
        ["Indicador", "Valor"],
        ["Pacientes clínicos", assistencial.get("total_pacientes_clinicos", 0)],
        ["Aferições de PA", assistencial.get("total_afericoes_pa", 0)],
        ["Aferições glicêmicas", assistencial.get("total_glicemias", 0)],
        ["Bioimpedâncias", assistencial.get("total_bioimpedancias", 0)],
        ["Intervenções farmacêuticas", assistencial.get("total_intervencoes", 0)],
    ])

    adicionar_tabela("Indicadores Cardiovasculares", [
        ["Indicador", "Valor"],
        ["PAS média", cardiovascular.get("media_pas", 0)],
        ["PAD média", cardiovascular.get("media_pad", 0)],
    ])

    adicionar_tabela("Indicadores Glicêmicos", [
        ["Indicador", "Valor"],
        ["Glicemia média", glicemico.get("media_glicemia", 0)],
    ])

    adicionar_tabela("Indicadores Antropométricos", [
        ["Indicador", "Valor"],
        ["IMC médio", antropometrico.get("media_imc", 0)],
    ])

    adicionar_tabela("Intervenções Farmacêuticas", [
        ["Indicador", "Valor"],
        ["Intervenções aceitas", intervencoes.get("intervencoes_aceitas", 0)],
        ["Encaminhamentos", intervencoes.get("encaminhamentos", 0)],
        ["Taxa de aceitação (%)", intervencoes.get("taxa_aceitacao", 0)],
        ["Taxa de encaminhamento (%)", intervencoes.get("taxa_encaminhamento", 0)],
    ])

    dados_serie = [[
        "Mês",
        "PA",
        "Glicemia",
        "Bioimpedância",
        "Intervenções"
    ]]

    for item in serie:
        dados_serie.append([
            item.get("mes", ""),
            item.get("afericoes_pa", 0),
            item.get("glicemias", 0),
            item.get("bioimpedancias", 0),
            item.get("intervencoes", 0),
        ])

    adicionar_tabela("Série Temporal Científica", dados_serie)

    elementos.append(Paragraph("Interpretação resumida", styles["Heading2"]))
    elementos.append(Paragraph(
        "Este relatório consolida indicadores assistenciais, epidemiológicos e clínicos "
        "registrados no sistema. Os dados têm finalidade de apoio à gestão, ensino, "
        "pesquisa, extensão e avaliação dos serviços farmacêuticos. A interpretação "
        "científica deve considerar o desenho observacional dos registros e a qualidade "
        "do preenchimento dos dados.",
        styles["Normal"]
    ))

    elementos.append(Spacer(1, 30))

    elementos.append(Paragraph(
        f"Profissional responsável: {getattr(current, 'nome', 'Não informado')}",
        styles["Normal"]
    ))

    elementos.append(Paragraph(
        f"Categoria profissional: {getattr(current, 'categoria_profissional', 'Não informado')}",
        styles["Normal"]
    ))

    doc.build(elementos)

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "inline; filename=relatorio_cientifico.pdf"
        }
    )
