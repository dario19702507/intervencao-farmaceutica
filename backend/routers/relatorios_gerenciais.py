"""Relatórios gerenciais da Farmácia Escola.

Passo 13B: consolida dados operacionais em relatórios JSON, CSV e PDF.
A primeira versão prioriza três eixos já maduros no sistema:
- operacional diário/período;
- vigências/laudos;
- gestão documental.
"""

from __future__ import annotations

import csv
import io
from datetime import date, datetime, timedelta
from typing import Iterable, Optional

from fastapi import APIRouter, Depends, Query, Response
from sqlalchemy.orm import Session
from sqlalchemy import func
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

from routers.consultorio import get_db_consultorio, get_current_user_consultorio
from services.documentos_institucionais import cabecalho_institucional, rodape_institucional, LOGO_FARMACIA, LOGO_UFMS

from models.consultorio_models import (
    AgendaIntegrada,
    ProcessoDocumental,
    DocumentoPaciente,
    NotificacaoInterna,
    WhatsAppEnvio,
)

router = APIRouter(prefix="/consultorio/relatorios-gerenciais", tags=["Relatórios Gerenciais"])


TIPOS_RELATORIO = ["OPERACIONAL", "VIGENCIAS", "DOCUMENTAL"]
FORMATOS_EXPORTACAO = ["JSON", "CSV", "XLSX", "PDF"]


def _hoje() -> date:
    return date.today()


def _parse_date(value: Optional[str], default: Optional[date] = None) -> Optional[date]:
    if not value:
        return default
    if isinstance(value, date):
        return value
    return datetime.strptime(value, "%Y-%m-%d").date()


def _periodo_padrao(data_inicio: Optional[str], data_fim: Optional[str]) -> tuple[date, date]:
    hoje = _hoje()
    inicio = _parse_date(data_inicio, hoje)
    fim = _parse_date(data_fim, hoje)
    return inicio, fim


def _contar(query) -> int:
    return int(query.count() or 0)


def _linha_agenda(e: AgendaIntegrada) -> dict:
    return {
        "id": e.id,
        "data": e.data_evento.isoformat() if e.data_evento else None,
        "tipo_evento": e.tipo_evento,
        "paciente": e.paciente_nome,
        "medicamento": e.medicamento,
        "prioridade": e.prioridade,
        "status": e.status,
        "titulo": e.titulo,
    }


def _linha_processo(p: ProcessoDocumental) -> dict:
    return {
        "id": p.id,
        "paciente_id": p.paciente_id,
        "tipo_processo": p.tipo_processo,
        "titulo": p.titulo,
        "situacao": p.situacao,
        "prioridade": p.prioridade,
        "data_abertura": p.data_abertura.isoformat() if p.data_abertura else None,
        "data_conclusao": p.data_conclusao.isoformat() if p.data_conclusao else None,
        "vigencia_inicio": p.vigencia_inicio.isoformat() if p.vigencia_inicio else None,
        "vigencia_fim": p.vigencia_fim.isoformat() if p.vigencia_fim else None,
        "vigencia_status": p.vigencia_status,
        "pendencias": p.pendencias_descricao,
    }


def _linha_documento(d: DocumentoPaciente) -> dict:
    return {
        "id": d.id,
        "paciente_id": d.paciente_id,
        "processo_documental_id": d.processo_documental_id,
        "tipo_documento": d.tipo_documento,
        "titulo": d.titulo,
        "arquivo": d.nome_arquivo_original,
        "data_emissao": d.data_emissao.isoformat() if d.data_emissao else None,
        "data_validade": d.data_validade.isoformat() if d.data_validade else None,
        "status_documental": getattr(d, "status_documental", None),
        "ativo": d.ativo,
        "criado_em": d.criado_em.isoformat() if d.criado_em else None,
    }


def gerar_relatorio_operacional(db: Session, data_inicio: date, data_fim: date) -> dict:
    agenda_base = db.query(AgendaIntegrada).filter(
        AgendaIntegrada.data_evento >= data_inicio,
        AgendaIntegrada.data_evento <= data_fim,
    )

    retiradas = agenda_base.filter(AgendaIntegrada.tipo_evento == "RETIRADA")
    renovacoes = agenda_base.filter(AgendaIntegrada.tipo_evento == "RENOVACAO")
    inclusoes = agenda_base.filter(AgendaIntegrada.tipo_evento == "INCLUSAO")
    adequacoes = agenda_base.filter(AgendaIntegrada.tipo_evento == "ADEQUACAO")

    pendentes = agenda_base.filter(AgendaIntegrada.status.in_(["AGENDADO", "agendado", "PENDENTE"]))
    realizados = agenda_base.filter(AgendaIntegrada.status.in_(["REALIZADO", "realizado", "CONCLUIDO"]))
    atrasados = db.query(AgendaIntegrada).filter(
        AgendaIntegrada.data_evento < _hoje(),
        AgendaIntegrada.status.in_(["AGENDADO", "agendado", "PENDENTE"]),
    )

    eventos = agenda_base.order_by(AgendaIntegrada.data_evento.asc(), AgendaIntegrada.prioridade.desc()).limit(100).all()

    return {
        "tipo": "OPERACIONAL",
        "periodo": {"inicio": data_inicio.isoformat(), "fim": data_fim.isoformat()},
        "indicadores": {
            "eventos_periodo": _contar(agenda_base),
            "retiradas_previstas": _contar(retiradas),
            "renovacoes_previstas": _contar(renovacoes),
            "inclusoes_previstas": _contar(inclusoes),
            "adequacoes_previstas": _contar(adequacoes),
            "eventos_pendentes": _contar(pendentes),
            "eventos_realizados": _contar(realizados),
            "eventos_atrasados": _contar(atrasados),
            "notificacoes_nao_lidas": _contar(db.query(NotificacaoInterna).filter(NotificacaoInterna.lida == False)),
            "whatsapp_pendente": _contar(db.query(WhatsAppEnvio).filter(WhatsAppEnvio.status == "PENDENTE")),
        },
        "eventos": [_linha_agenda(e) for e in eventos],
    }


def gerar_relatorio_vigencias(db: Session, data_referencia: date) -> dict:
    limite_60 = data_referencia + timedelta(days=60)

    processos_com_vigencia = db.query(ProcessoDocumental).filter(ProcessoDocumental.vigencia_fim.isnot(None))
    vencendo = processos_com_vigencia.filter(
        ProcessoDocumental.vigencia_fim >= data_referencia,
        ProcessoDocumental.vigencia_fim <= limite_60,
    )
    vencidos = processos_com_vigencia.filter(ProcessoDocumental.vigencia_fim < data_referencia)
    vigentes = processos_com_vigencia.filter(
        ProcessoDocumental.vigencia_inicio <= data_referencia,
        ProcessoDocumental.vigencia_fim >= data_referencia,
    )

    docs_vencendo = db.query(DocumentoPaciente).filter(
        DocumentoPaciente.ativo == True,
        DocumentoPaciente.data_validade.isnot(None),
        DocumentoPaciente.data_validade >= data_referencia,
        DocumentoPaciente.data_validade <= limite_60,
    )
    docs_vencidos = db.query(DocumentoPaciente).filter(
        DocumentoPaciente.ativo == True,
        DocumentoPaciente.data_validade.isnot(None),
        DocumentoPaciente.data_validade < data_referencia,
    )

    lista_vencendo = vencendo.order_by(ProcessoDocumental.vigencia_fim.asc()).limit(100).all()
    lista_vencidos = vencidos.order_by(ProcessoDocumental.vigencia_fim.asc()).limit(100).all()

    return {
        "tipo": "VIGENCIAS",
        "data_referencia": data_referencia.isoformat(),
        "janela_dias": 60,
        "indicadores": {
            "processos_vigentes": _contar(vigentes),
            "processos_vencendo_60_dias": _contar(vencendo),
            "processos_vencidos": _contar(vencidos),
            "documentos_vencendo_60_dias": _contar(docs_vencendo),
            "documentos_vencidos": _contar(docs_vencidos),
        },
        "processos_vencendo": [_linha_processo(p) for p in lista_vencendo],
        "processos_vencidos": [_linha_processo(p) for p in lista_vencidos],
    }


def gerar_relatorio_documental(db: Session, data_inicio: date, data_fim: date) -> dict:
    processos = db.query(ProcessoDocumental).filter(
        ProcessoDocumental.data_abertura >= data_inicio,
        ProcessoDocumental.data_abertura <= data_fim,
    )
    documentos = db.query(DocumentoPaciente).filter(DocumentoPaciente.ativo == True)

    completos = processos.filter(ProcessoDocumental.situacao == "COMPLETO")
    incompletos = processos.filter(ProcessoDocumental.situacao.in_(["INCOMPLETO", "EM_MONTAGEM", "AGUARDANDO_DOCUMENTOS"]))
    sem_documentos = processos.filter(ProcessoDocumental.situacao == "SEM_DOCUMENTOS")

    docs_recebidos = documentos.filter(DocumentoPaciente.status_documental == "RECEBIDO")
    docs_validados = documentos.filter(DocumentoPaciente.status_documental == "VALIDADO")
    docs_rejeitados = documentos.filter(DocumentoPaciente.status_documental == "REJEITADO")
    docs_substituidos = documentos.filter(DocumentoPaciente.status_documental == "SUBSTITUIDO")

    processos_criticos = incompletos.order_by(ProcessoDocumental.prioridade.desc(), ProcessoDocumental.data_abertura.asc()).limit(100).all()
    documentos_rejeitados = docs_rejeitados.order_by(DocumentoPaciente.status_documental_atualizado_em.desc().nullslast()).limit(100).all()

    return {
        "tipo": "DOCUMENTAL",
        "periodo": {"inicio": data_inicio.isoformat(), "fim": data_fim.isoformat()},
        "indicadores": {
            "processos_periodo": _contar(processos),
            "processos_completos": _contar(completos),
            "processos_incompletos": _contar(incompletos),
            "processos_sem_documentos": _contar(sem_documentos),
            "documentos_recebidos": _contar(docs_recebidos),
            "documentos_validados": _contar(docs_validados),
            "documentos_rejeitados": _contar(docs_rejeitados),
            "documentos_substituidos": _contar(docs_substituidos),
        },
        "processos_incompletos": [_linha_processo(p) for p in processos_criticos],
        "documentos_rejeitados": [_linha_documento(d) for d in documentos_rejeitados],
    }


def _flatten_relatorio(relatorio: dict) -> list[dict]:
    rows: list[dict] = []
    for chave, valor in relatorio.get("indicadores", {}).items():
        rows.append({"secao": "indicadores", "item": chave, "valor": valor})
    for secao in ["eventos", "processos_vencendo", "processos_vencidos", "processos_incompletos", "documentos_rejeitados"]:
        for item in relatorio.get(secao, []) or []:
            row = {"secao": secao}
            row.update(item)
            rows.append(row)
    return rows


def _formatar_cabecalho(campo: str) -> str:
    return str(campo).replace("_", " ").strip().title()


def _formatar_valor(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return "Sim" if valor else "Não"
    return str(valor)


def _colunas_por_secao(secao: str, rows: list[dict]) -> list[str]:
    preferenciais = {
        "indicadores": ["item", "valor"],
        "eventos": ["data", "tipo_evento", "paciente", "medicamento", "prioridade", "status", "titulo"],
        "processos_vencendo": ["id", "tipo_processo", "titulo", "situacao", "prioridade", "vigencia_inicio", "vigencia_fim", "vigencia_status"],
        "processos_vencidos": ["id", "tipo_processo", "titulo", "situacao", "prioridade", "vigencia_inicio", "vigencia_fim", "vigencia_status"],
        "processos_incompletos": ["id", "tipo_processo", "titulo", "situacao", "prioridade", "data_abertura", "pendencias"],
        "documentos_rejeitados": ["id", "tipo_documento", "titulo", "arquivo", "status_documental", "data_validade", "processo_documental_id"],
    }
    existentes = {k for row in rows for k in row.keys()}
    cols = [c for c in preferenciais.get(secao, []) if c in existentes]
    extras = sorted(existentes - set(cols) - {"secao"})
    return cols + extras[: max(0, 8 - len(cols))]


def _linhas_por_secao(relatorio: dict) -> dict[str, list[dict]]:
    secoes: dict[str, list[dict]] = {}
    indicadores = []
    for chave, valor in relatorio.get("indicadores", {}).items():
        indicadores.append({"item": chave.replace("_", " ").title(), "valor": valor})
    secoes["indicadores"] = indicadores

    for secao in ["eventos", "processos_vencendo", "processos_vencidos", "processos_incompletos", "documentos_rejeitados"]:
        itens = relatorio.get(secao, []) or []
        if itens:
            secoes[secao] = itens
    return secoes


def _csv_response(relatorio: dict, filename: str) -> Response:
    # Excel em configuração pt-BR geralmente abre CSV corretamente com separador ponto-e-vírgula.
    # O BOM ajuda o Excel a reconhecer UTF-8 e preservar acentos.
    secoes = _linhas_por_secao(relatorio)
    output = io.StringIO()
    for secao, rows in secoes.items():
        if not rows:
            continue
        output.write(f"{secao.replace('_', ' ').title()}\n")
        fieldnames = _colunas_por_secao(secao, rows)
        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore", delimiter=";")
        writer.writeheader()
        writer.writerows(rows)
        output.write("\n")

    if not output.getvalue().strip():
        output.write("Secao;Item;Valor\n")

    return Response(
        content="\ufeff" + output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )



def _sheet_name(secao: str, used: set[str]) -> str:
    nomes = {
        "indicadores": "Indicadores",
        "eventos": "Eventos",
        "processos_vencendo": "Processos Vencendo",
        "processos_vencidos": "Processos Vencidos",
        "processos_incompletos": "Processos Incompletos",
        "documentos_rejeitados": "Documentos Rejeitados",
    }
    base = nomes.get(secao, secao.replace("_", " ").title())[:31]
    name = base
    idx = 2
    while name in used:
        suffix = f" {idx}"
        name = f"{base[:31-len(suffix)]}{suffix}"
        idx += 1
    used.add(name)
    return name


def _xlsx_response(relatorio: dict, filename: str) -> Response:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage
    except Exception as exc:
        return Response(
            content=(
                "Dependência ausente para exportação XLSX. "
                "Instale com: pip install openpyxl"
            ),
            status_code=500,
            media_type="text/plain; charset=utf-8",
        )

    secoes = _linhas_por_secao(relatorio)
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=14, color="0F172A")
    subtitle_font = Font(size=10, color="475569")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    used_names: set[str] = set()

    for secao, rows in secoes.items():
        if not rows:
            continue
        ws = wb.create_sheet(_sheet_name(secao, used_names))
        ws.sheet_view.showGridLines = False
        try:
            if LOGO_FARMACIA.exists():
                img_fe = XLImage(str(LOGO_FARMACIA))
                img_fe.width = 120
                img_fe.height = 80
                ws.add_image(img_fe, "A1")
            if LOGO_UFMS.exists():
                img_ufms = XLImage(str(LOGO_UFMS))
                img_ufms.width = 180
                img_ufms.height = 65
                ws.add_image(img_ufms, "D1")
            ws.row_dimensions[1].height = 60
        except Exception:
            pass
        ws["A3"] = "Farmácia Escola UFMS"
        ws["A3"].font = title_font
        ws["A4"] = relatorio.get("tipo", "Relatório Gerencial")
        ws["A4"].font = subtitle_font
        ws["A5"] = f"Emitido em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A5"].font = subtitle_font

        if relatorio.get("periodo"):
            periodo = relatorio["periodo"]
            ws["A6"] = f"Período: {periodo.get('inicio', '')} a {periodo.get('fim', '')}"
        elif relatorio.get("data_referencia"):
            ws["A6"] = f"Data de referência: {relatorio.get('data_referencia')}"
        ws["A6"].font = subtitle_font

        fieldnames = _colunas_por_secao(secao, rows)
        start_row = 8
        for col_idx, campo in enumerate(fieldnames, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=_formatar_cabecalho(campo))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row_idx, row in enumerate(rows, start=start_row + 1):
            for col_idx, campo in enumerate(fieldnames, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=_formatar_valor(row.get(campo, "")))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

        ws.freeze_panes = "A7"
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(fieldnames))}{max(start_row + len(rows), start_row)}"

        for col_idx, campo in enumerate(fieldnames, start=1):
            letra = get_column_letter(col_idx)
            max_len = len(_formatar_cabecalho(campo))
            for row in rows[:200]:
                max_len = max(max_len, len(_formatar_valor(row.get(campo, ""))))
            ws.column_dimensions[letra].width = min(max(max_len + 2, 12), 42)
        ws.row_dimensions[start_row].height = 24

    if not wb.sheetnames:
        ws = wb.create_sheet("Relatorio")
        ws["A1"] = "Nenhum dado disponível"

    buffer = io.BytesIO()
    wb.save(buffer)
    xlsx = buffer.getvalue()
    buffer.close()
    return Response(
        content=xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

def _pdf_response(relatorio: dict, titulo: str, filename: str) -> Response:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    story = []

    story.extend(cabecalho_institucional(styles, titulo))
    story.append(Paragraph(f"Emitido em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    story.append(Spacer(1, 0.35 * cm))

    secoes = _linhas_por_secao(relatorio)
    for secao, rows in secoes.items():
        if not rows:
            continue
        story.append(Paragraph(secao.replace("_", " ").title(), styles["Heading3"]))
        fieldnames = _colunas_por_secao(secao, rows)
        cabecalho = [Paragraph(_formatar_cabecalho(c), styles["BodyText"]) for c in fieldnames]
        dados = [cabecalho]
        for row in rows[:80]:
            dados.append([Paragraph(_formatar_valor(row.get(c, ""))[:180], styles["BodyText"]) for c in fieldnames])

        page_width, _ = landscape(A4)
        usable_width = page_width - 2.4 * cm
        col_width = usable_width / max(len(fieldnames), 1)
        tabela = Table(dados, colWidths=[col_width] * len(fieldnames), repeatRows=1)
        tabela.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAEAEA")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(tabela)
        story.append(Spacer(1, 0.35 * cm))

    doc.build(story, onFirstPage=lambda canvas, doc: rodape_institucional(canvas, doc, None), onLaterPages=lambda canvas, doc: rodape_institucional(canvas, doc, None))
    pdf = buffer.getvalue()
    buffer.close()
    return Response(
        content=pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/opcoes")
def opcoes_relatorios_gerenciais(current_user=Depends(get_current_user_consultorio)):
    return {
        "tipos_relatorio": TIPOS_RELATORIO,
        "formatos_exportacao": FORMATOS_EXPORTACAO,
        "observacao": "Relatórios gerenciais consolidados para acompanhamento operacional da Farmácia Escola.",
    }


@router.get("/operacional")
def relatorio_operacional(
    data_inicio: Optional[str] = Query(None),
    data_fim: Optional[str] = Query(None),
    db: Session = Depends(get_db_consultorio),
    current_user=Depends(get_current_user_consultorio),
):
    inicio, fim = _periodo_padrao(data_inicio, data_fim)
    return gerar_relatorio_operacional(db, inicio, fim)


@router.get("/vigencias")
def relatorio_vigencias(
    data_referencia: Optional[str] = Query(None),
    db: Session = Depends(get_db_consultorio),
    current_user=Depends(get_current_user_consultorio),
):
    referencia = _parse_date(data_referencia, _hoje())
    return gerar_relatorio_vigencias(db, referencia)


@router.get("/documental")
def relatorio_documental(
    data_inicio: Optional[str] = Query(None),
    data_fim: Optional[str] = Query(None),
    db: Session = Depends(get_db_consultorio),
    current_user=Depends(get_current_user_consultorio),
):
    inicio, fim = _periodo_padrao(data_inicio, data_fim)
    return gerar_relatorio_documental(db, inicio, fim)


@router.get("/operacional/csv")
def relatorio_operacional_csv(data_inicio: Optional[str] = None, data_fim: Optional[str] = None, db: Session = Depends(get_db_consultorio), current_user=Depends(get_current_user_consultorio)):
    inicio, fim = _periodo_padrao(data_inicio, data_fim)
    return _csv_response(gerar_relatorio_operacional(db, inicio, fim), "relatorio_operacional.csv")


@router.get("/vigencias/csv")
def relatorio_vigencias_csv(data_referencia: Optional[str] = None, db: Session = Depends(get_db_consultorio), current_user=Depends(get_current_user_consultorio)):
    referencia = _parse_date(data_referencia, _hoje())
    return _csv_response(gerar_relatorio_vigencias(db, referencia), "relatorio_vigencias.csv")


@router.get("/documental/csv")
def relatorio_documental_csv(data_inicio: Optional[str] = None, data_fim: Optional[str] = None, db: Session = Depends(get_db_consultorio), current_user=Depends(get_current_user_consultorio)):
    inicio, fim = _periodo_padrao(data_inicio, data_fim)
    return _csv_response(gerar_relatorio_documental(db, inicio, fim), "relatorio_documental.csv")


@router.get("/operacional/xlsx")
def relatorio_operacional_xlsx(data_inicio: Optional[str] = None, data_fim: Optional[str] = None, db: Session = Depends(get_db_consultorio), current_user=Depends(get_current_user_consultorio)):
    inicio, fim = _periodo_padrao(data_inicio, data_fim)
    return _xlsx_response(gerar_relatorio_operacional(db, inicio, fim), "relatorio_operacional.xlsx")


@router.get("/vigencias/xlsx")
def relatorio_vigencias_xlsx(data_referencia: Optional[str] = None, db: Session = Depends(get_db_consultorio), current_user=Depends(get_current_user_consultorio)):
    referencia = _parse_date(data_referencia, _hoje())
    return _xlsx_response(gerar_relatorio_vigencias(db, referencia), "relatorio_vigencias.xlsx")


@router.get("/documental/xlsx")
def relatorio_documental_xlsx(data_inicio: Optional[str] = None, data_fim: Optional[str] = None, db: Session = Depends(get_db_consultorio), current_user=Depends(get_current_user_consultorio)):
    inicio, fim = _periodo_padrao(data_inicio, data_fim)
    return _xlsx_response(gerar_relatorio_documental(db, inicio, fim), "relatorio_documental.xlsx")


@router.get("/operacional/pdf")
def relatorio_operacional_pdf(data_inicio: Optional[str] = None, data_fim: Optional[str] = None, db: Session = Depends(get_db_consultorio), current_user=Depends(get_current_user_consultorio)):
    inicio, fim = _periodo_padrao(data_inicio, data_fim)
    return _pdf_response(gerar_relatorio_operacional(db, inicio, fim), "Relatório Operacional", "relatorio_operacional.pdf")


@router.get("/vigencias/pdf")
def relatorio_vigencias_pdf(data_referencia: Optional[str] = None, db: Session = Depends(get_db_consultorio), current_user=Depends(get_current_user_consultorio)):
    referencia = _parse_date(data_referencia, _hoje())
    return _pdf_response(gerar_relatorio_vigencias(db, referencia), "Relatório de Vigências", "relatorio_vigencias.pdf")


@router.get("/documental/pdf")
def relatorio_documental_pdf(data_inicio: Optional[str] = None, data_fim: Optional[str] = None, db: Session = Depends(get_db_consultorio), current_user=Depends(get_current_user_consultorio)):
    inicio, fim = _periodo_padrao(data_inicio, data_fim)
    return _pdf_response(gerar_relatorio_documental(db, inicio, fim), "Relatório Documental", "relatorio_documental.pdf")
